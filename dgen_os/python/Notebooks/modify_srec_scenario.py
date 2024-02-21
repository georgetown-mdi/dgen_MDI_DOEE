import argparse
import psycopg2.extras as pgx
import psycopg2 as pg
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import time
import os
import sqlalchemy
from sqlalchemy import create_engine
import pyarrow as pa
import pyarrow.parquet as pq
from IPython.display import HTML
import base64
from openpyxl import load_workbook


from pylab import *
import geopandas as gpd
import pickle
import zipfile


#########################################################################

# set SREC

def modify_sql_line(file_path, marker, new_value):
    # Temp file to store changes
    temp_file_path = file_path + '.tmp'

    with open(file_path, 'r') as file, open(temp_file_path, 'w') as temp_file:
        for line in file:
            if line.startswith(marker):
                parts = line.split('\t')

                # Assuming the value to change is the 7th element (index 6)
                parts[6] = str(new_value)

                # Write the modified line
                temp_file.write('\t'.join(parts))
            else:
                # Write unmodified lines
                temp_file.write(line)

    # Replace the original file with the modified one
    os.replace(temp_file_path, file_path)


def main(args):
    # Modify SQL Line
    modify_sql_line(args.database_path, args.marker, args.srec)

    # Set Scenario
    # Map of parameters to their new values as provided
    parameters_to_update = {
        "Technology": "Solar + Storage",
        "Agent File": "Use pre-generated Agents",
        "Region to Analyze": "District of Columbia",
        "Markets": "Only Residential",
        "Analysis End Year": int(2040),
        "Load Growth Scenario": "AEO2019 High Price", # options above
        "Random Generator Seed": 22      
    }

    # Map of scenario names to their new values as provided
    scenario_values = {
        "Retail Electricity Price Escalation Scenario": "ATB19_High_RE_Cost_retail",
        "Wholesale Electricity Price Scenario": "ATB19_Mid_Case_wholesale",
        "PV Price Scenario": "pv_price_atb19_high", #pv_price_atb19_mid, pv_price_atb19_low
        "PV Technical Performance Scenario": "pv_tech_performance_defaultFY19",
        "Storage Cost Scenario": "batt_prices_FY20_low", #batt_prices_FY20_mid
        "Storage Technical Performance Scenario": "batt_tech_performance_SunLamp17",
        "PV + Storage Cost Scenario": "pv_plus_batt_prices_FY20_mid",
        "Financing Scenario": "financing_atb_FY19",
        "Depreciation Scenario": "deprec_sch_FY19",
        "Value of Resiliency Scenario": "vor_FY20_mid",
        "Carbon Intensity Scenario": "carbon_intensities_FY19"
    }

    # Load the workbook and select the active worksheet
    input_sheet_final = os.path.join(args.input_sheet_path, "input_sheet_final.xlsx")
    workbook = load_workbook(filename=input_sheet_final)    
    sheet = workbook.active

    # PARAMETER SET 1
    for param, value in parameters_to_update.items():
        # Find the row index where the parameter name matches in the 3rd column (index 2)
        for row in sheet.iter_rows(min_row=1, max_col=3, max_row=sheet.max_row):
            for cell in row:
                if cell.value == param:
                    sheet.cell(row=cell.row, column=cell.column+1).value = value
                    break

    # PARAMETER SET 2                
    for param, value in scenario_values.items():
        # Find the row index where the parameter name matches in the 3rd column (index 2)
        for row in sheet.iter_rows(min_row=1, max_col=3, max_row=sheet.max_row):
            for cell in row:
                if cell.value == param:
                    sheet.cell(row=cell.row, column=cell.column+2).value = value
                    break
                   
    # Save the workbook 
    workbook.save(filename=input_sheet_final)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Modify SQL line and set scenario parameters")
    parser.add_argument("--srec", type=float, help="New SREC value", required=True)
    parser.add_argument("--input_sheet_path", type=str, help="Path to input sheet directory", required=True)
    parser.add_argument("--database_path", type=str, help="Path to database directory", required=True)
    parser.add_argument("--marker", type=str, help="Marker string for SQL line modification", default="DC\tres\tsolar\tSREC\t\\N\t\\N\t")
    args = parser.parse_args()
    main(args)
